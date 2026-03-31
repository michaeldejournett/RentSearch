"""
Assembles the results DataFrame and exports a styled Excel workbook in memory.
"""

import io
import re
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .geocode import compute_weighted_distance, compute_distance_score, is_too_far


def _parse_price(raw) -> Optional[float]:
    """Normalise a price value (string or number) to a float."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    cleaned = re.sub(r"[^\d.]", "", str(raw))
    try:
        return float(cleaned)
    except ValueError:
        return None


def compute_total_score(
    criteria_scores: list[Optional[float]],
    criteria_weights: list[int],
    distance_score: float,
    distance_weight: float = 5.0,
) -> float:
    """Weighted combination of criteria + distance scores, result 0–10."""
    valid_pairs = [
        (s, w) for s, w in zip(criteria_scores, criteria_weights) if s is not None
    ]
    c_weight_total = sum(w for _, w in valid_pairs)
    c_score = (
        sum(s * w for s, w in valid_pairs) / c_weight_total if c_weight_total > 0 else 0.0
    )
    total_weight = c_weight_total + distance_weight
    if total_weight == 0:
        return 0.0
    total = (c_score * c_weight_total + distance_score * distance_weight) / total_weight
    return round(total, 2)


def _score_color(score: float) -> str:
    """Return an ARGB hex color for a 0–10 score (red → yellow → green)."""
    ratio = max(0.0, min(1.0, score / 10.0))
    if ratio < 0.5:
        r, g = 255, int(255 * ratio * 2)
    else:
        r, g = int(255 * (1 - ratio) * 2), 255
    return f"FF{r:02X}{g:02X}00"


def build_dataframe(
    listings: list[dict],
    locations: list[dict],
    criteria: list[dict],
    min_baths: Optional[float] = None,
    max_baths: Optional[float] = None,
    min_beds: Optional[int] = None,
    max_beds: Optional[int] = None,
    listing_names: Optional[dict] = None,
) -> pd.DataFrame:
    """Build the ranked results DataFrame."""
    rows = []

    for listing in listings:
        ext = listing.get("extracted") or {}
        scoring = listing.get("scoring") or {}

        # --- minimum data check — skip listings with no extraction data at all ---
        price_raw = _parse_price(ext.get("price_monthly"))
        bedrooms_raw = ext.get("bedrooms")
        address_raw = ext.get("address")
        summary_raw = scoring.get("overall_summary", "")
        has_structured = price_raw is not None or bedrooms_raw is not None or address_raw
        has_summary = bool(summary_raw and summary_raw.strip())
        if not has_structured and not has_summary:
            continue

        address = address_raw or "—"
        apt_coords = listing.get("apt_coords")

        # Distance: hard-filter if coords available and listing exceeds every location's max
        if apt_coords and locations:
            if is_too_far(apt_coords, locations):
                continue
            dist_info = compute_weighted_distance(apt_coords, locations)
            per_loc = dist_info["per_location"]
            dist_score = compute_distance_score(apt_coords, locations)
        else:
            per_loc = {}
            dist_score = 0.0

        # --- bedrooms filter ---
        bedrooms = bedrooms_raw  # already extracted above
        if bedrooms is not None:
            if min_beds is not None and bedrooms < min_beds:
                continue
            if max_beds is not None and bedrooms > max_beds:
                continue

        # --- bathrooms filter ---
        bathrooms = ext.get("bathrooms")
        if bathrooms is not None:
            if min_baths is not None and bathrooms < min_baths:
                continue
            if max_baths is not None and bathrooms > max_baths:
                continue

        # --- criteria scores ---
        score_entries = scoring.get("scores", [])
        # Build a map: criterion text → {score, note}
        score_map: dict[str, dict] = {}
        for entry in score_entries:
            key = entry.get("criterion", "")
            score_map[key] = {"score": entry.get("score"), "note": entry.get("note", "")}

        criteria_scores = []
        criteria_weights = []
        for c in criteria:
            entry = score_map.get(c["text"], {})
            criteria_scores.append(entry.get("score"))
            criteria_weights.append(c["weight"])

        total_score = compute_total_score(
            criteria_scores, criteria_weights, dist_score
        )

        # --- build row ---
        href = listing.get("href", "")
        row: dict = {
            "Name": (listing_names or {}).get(href, ""),
            "Address": address,
            "Monthly Price ($)": price_raw,
            "Bedrooms": bedrooms_raw,
            "Bathrooms": ext.get("bathrooms"),
            "Sq Ft": ext.get("sqft"),
            "Distance Score": dist_score,
        }

        # Distance per location columns
        valid_locations = [loc for loc in locations if loc.get("coords")]
        for loc in valid_locations:
            label = loc.get("label", "Location")
            row[f"distance_to_{label}"] = per_loc.get(label)

        # Criteria score + note columns
        for c, score_val in zip(criteria, criteria_scores):
            key = c["text"]
            entry = score_map.get(key, {})
            row[f'{key[:30]} Score'] = score_val
            row[f'{key[:30]} Note'] = entry.get("note", "")

        row["Total Score"] = total_score
        row["Available Date"] = ext.get("available_date")
        row["URL"] = listing.get("href", "")
        row["AI Summary"] = scoring.get("overall_summary", "")
        row["Data Confidence"] = ext.get("extraction_confidence", "unknown")

        rows.append(row)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df.sort_values("Total Score", ascending=False, inplace=True)
    df.insert(0, "Rank", range(1, len(df) + 1))
    df.reset_index(drop=True, inplace=True)
    return df


def _style_worksheet(ws, df: pd.DataFrame) -> None:
    """Apply Excel formatting: bold header, alternating rows, score colors, freeze."""
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Find "Total Score" column index (1-based)
    score_col_idx = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "Total Score":
            score_col_idx = col_idx
            break

    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=False, vertical="center")

            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                # Color Total Score cell
                if score_col_idx and cell.column == score_col_idx and cell.value is not None:
                    try:
                        cell.fill = PatternFill("solid", fgColor=_score_color(float(cell.value)))
                        cell.font = Font(bold=True)
                    except (ValueError, TypeError):
                        pass

    # Auto-fit column widths (capped at 50)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:  # noqa: BLE001
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Wrap AI Summary column
    for col_idx, cell in enumerate(ws[1], 1):
        if "Summary" in str(cell.value or ""):
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for c in row:
                    c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[get_column_letter(col_idx)].width = 40
            break


def export_to_excel(
    listings: list[dict],
    locations: list[dict],
    criteria: list[dict],
    min_baths: Optional[float] = None,
    max_baths: Optional[float] = None,
    min_beds: Optional[int] = None,
    max_beds: Optional[int] = None,
    listing_names: Optional[dict] = None,
) -> bytes:
    """Build the DataFrame, create a styled Excel workbook in memory, return bytes."""
    df = build_dataframe(listings, locations, criteria,
                         min_baths=min_baths, max_baths=max_baths,
                         min_beds=min_beds, max_beds=max_beds,
                         listing_names=listing_names)

    output = io.BytesIO()
    if df.empty:
        # Return a workbook with a single message
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws["A1"] = "No results found matching your criteria."
        wb.save(output)
        return output.getvalue()

    wb = Workbook()
    ws = wb.active
    ws.title = "Apartment Rankings"

    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _style_worksheet(ws, df)
    wb.save(output)
    return output.getvalue()
